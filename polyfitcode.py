from matplotlib import pyplot as plt
import pandas as pd
import numpy as np
import os
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook

data_files= os.listdir('C:/Users/files') #listing the dataset
workbook = xlsxwriter.Workbook(r'C:/Users/miss_/output.xlsx')
workbook.close()

#create a function to loop over the data set list and output the equations in excel files
def load_files(filenames):
    for filename in filenames:
        dframe = pd.read_csv(filename) #read data set from csv file
        df= pd.DataFrame(dframe)
        df.columns=["x","y"]
        deg =int(input("enter PolyFit degree= "))
        x=df.x
        y=df.y
        polyyfit= np.polyfit(x , y, deg, rcond=None, full=False) #input dataset (x,y) to the polyfit function 
        plt.plot(x,y,'o') #plot the the existing data set (x,y)
        plottfit=np.poly1d(polyyfit) 
        plt.plot(x , plottfit(x))
        plt.show()
        print(polyyfit)
        res=pd.DataFrame(polyyfit)
        writer = pd.ExcelWriter('output.xlsx',engine='openpyxl', mode='a') #define the excel file name and writer mode
        writer.book = load_workbook('output.xlsx')                         #load an existing excel file to add sheets in the next step
        res.to_excel(writer,sheet_name=filename,index=False)                           #add the equation parameters to the excel sheet
        writer.save()
        writer.close()
    return polyyfit  

load_files(data_files)